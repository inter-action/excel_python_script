from openpyxl import load_workbook
from openpyxl import utils
import glob
import os

import model

def get_working_sheet(workbook, sheet_title):
    result = list(filter(lambda x: sheet_title in x, workbook.sheetnames)) 

    if len(result) >= 1:
        sheet_title = result[0]
        sheet = workbook[sheet_title]
        return sheet
    else:
        print(f'sheetnames: {workbook.sheetnames}')
        return None

def get_workbook(filename, data_only):
    workbook = load_workbook(filename=filename, data_only=data_only)
    return workbook

def range_to_loc(range):
    min_col, min_row, max_col, max_row = utils.range_boundaries(range)
    return model.Loc(min_row, max_row, min_col, max_col)


def iter_cols_by_range(sheet, range: str):
    loc = range_to_loc(range)
    return sheet.iter_cols(loc.min_column, loc.max_column, loc.min_row, loc.max_row)



def excel_files(path):
    path = "{}/**/*.xlsx".format(path)
    excels = glob.glob(path)
    for excel_path in excels:
        excel_filename = os.path.basename(excel_path)
        if excel_filename.find("应收账款") != -1 and not excel_filename.startswith('~$'):
            yield [excel_path, excel_filename]
